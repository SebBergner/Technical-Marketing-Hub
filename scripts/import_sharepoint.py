#!/usr/bin/env python
"""Turn a SharePoint export into data/seed/assets.json.

Replaces the 28 hand-extracted mockup rows with the real Demo Catalog.

The structural rule, established by measuring the site rather than guessing:

    An asset is a TOP-LEVEL folder in Demo Catalog that carries a Demo Type.
    Everything beneath it is a resource.

All 452 such folders sit at depth 0, so resolving a file's owner is just its
first path segment — no heuristics, no prefix search.

Deliberately NOT imported: the Virtual Machine Catalog. Its 49 VMs are all 2024
or older and were judged out of scope. Its field mapping is recorded in
docs/ARCHITECTURE.md so the analysis is not lost, and this module's shape is the
seam to add it behind — a second source_system, nothing else changes.

Usage, from the repo root:
    python scripts/import_sharepoint.py            # writes data/seed/assets.json
    python scripts/import_sharepoint.py --dry-run  # report only
"""
from __future__ import annotations

import argparse
import collections
import json
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW = r"C:\Users\liwchen\OneDrive - PTC\Desktop\Sharepoint Raw Data.xlsx"
CATALOG = r"C:\Work\TDD Hub\SharepoinDemoCatalog.xlsx"
OUT = os.path.join(ROOT, "data", "seed", "assets.json")
LIBRARY = "sites/EXT-TDD/Demo Catalog"

# Field logic lives in backend.services.sharepoint_mapping so this importer and
# the live Graph sync cannot drift apart — they must produce identical assets
# from the same SharePoint content.
from backend.services import sharepoint_mapping as m       # noqa: E402

clean_text = m.clean_text
parse_lookup = m.parse_lookup
parse_segment = m.parse_segment
slugify = m.slugify
display_title = m.display_title
classify = m.classify
audience_of = m.audience_of
audio_of = m.audio_of
pick_main_video = m.pick_main_video
TYPE_MAP = m.TYPE_MAP


def load_sheet(path: str):
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(h) if h is not None else "" for h in rows[0]]
    return header, [r for r in rows[1:] if any(r)]


def build_assets(rows, col, descriptions) -> dict[str, dict]:
    assets: dict[str, dict] = {}
    taken: set[str] = set()

    def get(row, name):
        return row[col[name]] if name in col else None

    for row in rows:
        if str(get(row, "Item Type")) != "Folder":
            continue
        if str(get(row, "Path")) != LIBRARY or not get(row, "Demo Type"):
            continue

        name = clean_text(get(row, "Name"))
        demo_type = clean_text(get(row, "Demo Type"))
        if not name or demo_type not in TYPE_MAP:
            continue

        segment, all_segments = parse_segment(get(row, "Segment"))
        modified = get(row, "Modified")
        depth = m.content_depth(name)

        assets[name] = {
            "id": slugify(name, taken),
            "type": TYPE_MAP[demo_type],
            "title": display_title(name),
            "description": descriptions.get(name) or clean_text(get(row, "Description")),
            "products": parse_lookup(get(row, "Product")),
            "funnel_stage": None,
            "content_depth": depth,
            "language": m.parse_language(get(row, "Language")),
            "segment": segment,
            "industry": None,
            "value_drivers": [],
            "customer_facing": True,
            "has_narrated_audio": None,
            "named_customer": None,
            "uploaded_at": str(modified)[:10] if modified else None,
            "duration_seconds": None,
            "thumbnail_url": None,
            "source_item_id": f"{LIBRARY}/{name}",
            "brightcove_id": None,
            "consensus_uuid": None,
            "web_url": None,
            "rails": all_segments[1:],       # extra segments become rail membership
            "is_editor_pick": False,
            "value_roadmap": None,
            "stats": {"views": 0, "downloads": 0, "launches": 0, "shares": 0},
            "resources": [],
            "main_video": None,
            "resource_count": 0,
            "video_count": 0,
            "resource_counts": {},
        }
    return assets


def attach_resources(rows, col, assets) -> int:
    """Attribute every file to its top-level folder. Returns the orphan count."""
    def get(row, name):
        return row[col[name]] if name in col else None

    orphans = 0
    for row in rows:
        if str(get(row, "Item Type")) == "Folder":
            continue
        path = str(get(row, "Path") or "")
        if not path.startswith(LIBRARY):
            continue
        rest = path[len(LIBRARY):].strip("/")
        if not rest:
            continue
        folder, _, subfolder = rest.partition("/")
        owner = assets.get(folder)
        if owner is None:
            orphans += 1          # lives under a folder carrying no Demo Type
            continue
        filename = clean_text(get(row, "Name")) or ""
        owner["resources"].append(m.build_resource(filename, subfolder))
    return orphans


#: Kinds a person would actually pick from a list. CAD is excluded because 42%
#: of the catalogue is Creo version files (`part.prt.1`, `asm.asm.2`) — listing
#: them individually is noise, and it tripled the seed file.
BROWSABLE = {"video", "document", "dataset", "image"}


def derive_video_facts(assets) -> None:
    for asset in assets.values():
        m.derive_video_facts(asset)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--raw", default=RAW)
    parser.add_argument("--catalog", default=CATALOG)
    args = parser.parse_args()

    for path in (args.raw, args.catalog):
        if not os.path.exists(path):
            print(f"missing export: {path}")
            return 1

    header, rows = load_sheet(args.raw)
    col = {h: i for i, h in enumerate(header)}

    # The folder-level export carries a better Description: `Description2`
    # preserves newlines and has the highest coverage (93% vs 81%).
    cat_header, cat_rows = load_sheet(args.catalog)
    cat_col = {h: i for i, h in enumerate(cat_header)}
    desc_field = "Description2" if "Description2" in cat_col else "Description"
    descriptions = {}
    for row in cat_rows:
        name = clean_text(row[cat_col["Name"]])
        if name:
            descriptions[name] = clean_text(row[cat_col[desc_field]])

    assets = build_assets(rows, col, descriptions)
    orphans = attach_resources(rows, col, assets)
    derive_video_facts(assets)

    records = sorted(assets.values(), key=lambda a: a["title"].lower())

    by_type = collections.Counter(a["type"] for a in records)
    with_video = sum(1 for a in records if a["video_count"])
    with_main = sum(1 for a in records if a["main_video"])
    total_res = sum(a["resource_count"] for a in records)

    print(f"assets              {len(records)}   {dict(by_type)}")
    print(f"resources           {total_res}")
    print(f"  videos            {sum(a['video_count'] for a in records)}")
    print(f"  orphan files      {orphans}  (under folders carrying no Demo Type)")
    print()
    print(f"with >=1 video      {with_video}/{len(records)}")
    print(f"main video resolved {with_main}/{with_video}"
          f"  ({100 * with_main // max(1, with_video)}%) — the rest need a human")
    print(f"customer_facing     {sum(1 for a in records if a['customer_facing'])}")
    print(f"internal only       {sum(1 for a in records if not a['customer_facing'])}")
    print(f"audio stated        {sum(1 for a in records if a['has_narrated_audio'] is not None)}")
    print(f"content_depth set   {sum(1 for a in records if a['content_depth'])}"
          f"  (only where the name states it)")
    print(f"consensus_uuid      0  — absent from SharePoint entirely")
    print()
    for field in ("segment", "language", "products", "description"):
        print(f"  {field:<14} {sum(1 for a in records if a[field])}/{len(records)}")

    if args.dry_run:
        print("\n--dry-run: nothing written")
        return 0

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8", newline="\n") as handle:
        json.dump(records, handle, indent=2, ensure_ascii=False)
        handle.write("\n")
    print(f"\nwrote {len(records)} assets -> {os.path.relpath(OUT, ROOT)}"
          f"  ({os.path.getsize(OUT) / 1024:.0f} KB)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
